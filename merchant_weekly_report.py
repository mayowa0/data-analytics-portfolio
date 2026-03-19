"""
Automated Weekly Merchant Loan Report
--------------------------------------
Queries a PostgreSQL database every Friday at 12 PM,
calculates merchant loan disbursement statistics,
compares against the previous week, and sends a
formatted HTML email with an Excel attachment.

Author : Mayowa Alamutu
Stack  : Python · psycopg2 · pandas · openpyxl · ZeptoMail API
"""

import psycopg2
import pandas as pd
import requests
import base64
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# CONFIGURATION  (all secrets loaded from .env — never hardcoded)
# ---------------------------------------------------------------------------
load_dotenv()

DB_HOST = os.getenv('DB_HOST')        # e.g. "db.example.com"
DB_NAME = os.getenv('DB_NAME')        # e.g. "fintech_prod"
DB_USER = os.getenv('DB_USER')
DB_PASS = os.getenv('DB_PASS')
DB_PORT = os.getenv('DB_PORT', '5432')

EMAIL_API_KEY    = os.getenv('EMAIL_ZEPTOMAIL_API_KEY')
EMAIL_API_URL    = os.getenv('EMAIL_ZEPTOMAIL_API_URL')
EMAIL_SENDER     = os.getenv('EMAIL_SENDER')
EMAIL_RECIPIENTS = os.getenv('EMAIL_RECIPIENTS').split(',')
EMAIL_CC         = os.getenv('EMAIL_CC_RECIPIENTS', '').split(',') if os.getenv('EMAIL_CC_RECIPIENTS') else []


# ---------------------------------------------------------------------------
# 1. SQL QUERY
# ---------------------------------------------------------------------------

def get_report_query(start_date, end_date):
    """
    Returns merchant-level loan disbursement data for the given date window.
    Joins merchants → merchant_user_loans → personal_loans.
    Only includes approved/ongoing loans disbursed in the period.
    """
    return f"""
    SELECT
        m.business_name                                          AS "Merchant Name",
        MAX(CASE
            WHEN m.created_at >= '{start_date}' AND m.created_at < '{end_date}'
            THEN 'YES' ELSE 'NO' END)                           AS "Is Newly Onboarded",

        COUNT(CASE WHEN pl.status = 'approved'         THEN pl.loan_id END)  AS "Approved Loan Count",
        COUNT(CASE WHEN pl.status = 'ongoing'          THEN pl.loan_id END)  AS "Ongoing Loan Count",

        COUNT(CASE WHEN pl.is_created_by_admin = TRUE  THEN pl.loan_id END)  AS "Admin - Loan Count",
        COALESCE(SUM(CASE  WHEN pl.is_created_by_admin = TRUE
                           THEN pl.amount_requested END), 0)                 AS "Admin - Volume",

        COUNT(CASE WHEN pl.is_created_by_admin = FALSE THEN pl.loan_id END)  AS "App - Loan Count",
        COALESCE(SUM(CASE  WHEN pl.is_created_by_admin = FALSE
                           THEN pl.amount_requested END), 0)                 AS "App - Volume",

        COUNT(pl.loan_id)        AS "Total Loan Count",
        SUM(pl.amount_requested) AS "Total Loan Volume"

    FROM public.merchants m
    LEFT JOIN public.merchant_user_loans mul ON m.merchant_id = mul.merchant_id
    LEFT JOIN public.personal_loans      pl  ON mul.loan_id   = pl.loan_id

    WHERE pl.status IN ('approved', 'ongoing')
      AND pl.loan_disbursed_at >= '{start_date}'
      AND pl.loan_disbursed_at <  '{end_date}'
    GROUP BY m.business_name
    HAVING COUNT(pl.loan_id) > 0
    ORDER BY 9 DESC;
    """


# ---------------------------------------------------------------------------
# 2. SUMMARY CALCULATION
# ---------------------------------------------------------------------------

def pct_change(current, previous):
    if previous == 0:
        return 100.0 if current > 0 else 0.0
    return ((current - previous) / previous) * 100


def build_summary(df):
    """Aggregate raw query results into a flat summary dict."""
    if df.empty:
        return {k: 0 if not isinstance(v, list) else v
                for k, v in {
                    'merchants_onboarded': 0, 'onboarded_names': [],
                    'merchants_approved':  0, 'approved_names':  [],
                    'merchants_with_loans':0, 'loan_merchant_names': [],
                    'total_loans': 0,  'total_volume': 0.0,
                    'admin_loans': 0,  'admin_volume': 0.0,
                    'app_loans':   0,  'app_volume':   0.0,
                }.items()}

    onboarded  = df[df['Is Newly Onboarded'] == 'YES']
    with_loans = df[df['Total Loan Count']   > 0]

    return {
        'merchants_onboarded':  len(onboarded),
        'onboarded_names':      onboarded['Merchant Name'].tolist(),
        'merchants_approved':   int((df['Approved Loan Count'] > 0).sum()),
        'approved_names':       df[df['Approved Loan Count'] > 0]['Merchant Name'].tolist(),
        'merchants_with_loans': len(with_loans),
        'loan_merchant_names':  with_loans['Merchant Name'].tolist(),
        'total_loans':          int(df['Total Loan Count'].sum()),
        'total_volume':         float(df['Total Loan Volume'].sum()),
        'admin_loans':          int(df['Admin - Loan Count'].sum()),
        'admin_volume':         float(df['Admin - Volume'].sum()),
        'app_loans':            int(df['App - Loan Count'].sum()),
        'app_volume':           float(df['App - Volume'].sum()),
    }


# ---------------------------------------------------------------------------
# 3. EMAIL HTML BUILDER  (Outlook-safe table layout)
# ---------------------------------------------------------------------------

def _stat(label, value):
    return (f'<p style="margin:4px 0;font-size:15px;font-family:Arial,sans-serif;">'
            f'<strong style="color:#1A201D;">{label}</strong>'
            f'<strong style="color:#2A8851;">&nbsp;{value}</strong></p>')

def _bullet(text):
    return (f'<p style="margin:2px 0 2px 16px;font-size:13px;'
            f'font-family:Arial,sans-serif;color:#333;">&#8226;&nbsp;{text}</p>')

def _divider():
    return '<hr style="border:none;border-top:1px solid #c2d9bb;margin:8px 0;">'

def _ydivider():
    return '<hr style="border:none;border-top:1px solid #e8d87a;margin:8px 0;">'

def _comp(title, this_val, last_val, curr, prev, is_currency=False):
    change = curr - prev
    pct    = pct_change(curr, prev)
    color  = "#2A8851" if change >= 0 else "#E46159"
    lbl    = "Growth"  if change >= 0 else "Decline"
    fmt    = f"&#8358;{change:+,.2f}" if is_currency else f"{change:+,}"
    return (f'<p style="margin:4px 0;font-size:14px;font-family:Arial,sans-serif;">'
            f'<strong>{title}</strong><br>'
            f'This Week: <strong>{this_val}</strong>&nbsp;|&nbsp;Last Week: {last_val}<br>'
            f'{lbl}: <strong style="color:{color};">{fmt} ({pct:+.1f}%)</strong></p>')

def build_email_html(cs, ps, start_date, end_date):
    period = (f"{start_date.strftime('%Y-%m-%d (%a)')} to "
              f"{(end_date - timedelta(days=1)).strftime('%Y-%m-%d (%a)')}")
    ts     = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Build merchant name bullets
    def name_bullets(names):
        return ''.join(_bullet(f"{i+1}. {n}") for i, n in enumerate(names))

    onboard_html = (_stat("Total Merchants Onboarded:", cs['merchants_onboarded']) +
                    name_bullets(cs['onboarded_names']))
    approved_html = (_stat("Total Merchants With Approval:", cs['merchants_approved']) +
                     name_bullets(cs['approved_names']))
    loans_html    = (_stat("Total Merchants With Loans:", cs['merchants_with_loans']) +
                     name_bullets(cs['loan_merchant_names']))

    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f5f5f5;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f5f5f5;">
<tr><td align="center" style="padding:20px 10px;">
<table width="620" cellpadding="0" cellspacing="0" border="0" style="background:#fff;max-width:620px;">

  <!-- HEADER -->
  <tr><td align="center" bgcolor="#235D3B" style="background:#235D3B;padding:24px 20px;">
    <p style="margin:0 0 6px;font-size:22px;font-weight:bold;color:#fff;">Weekly Merchant Loan Report</p>
    <p style="margin:0;font-size:13px;color:#ccffcc;">Period: {period}</p>
  </td></tr>

  <!-- BODY -->
  <tr><td style="padding:24px 20px;">
    <p style="margin:0 0 10px;font-size:17px;font-weight:bold;color:#235D3B;
               border-bottom:2px solid #D9ECD4;padding-bottom:4px;">Summary</p>

    <div style="background:#D9ECD4;border-left:4px solid #2A8851;padding:16px;">
      {onboard_html}{_divider()}
      {approved_html}{_divider()}
      {loans_html}{_divider()}
      {_stat("Total Loans Disbursed:", cs['total_loans'])}{_divider()}
      {_stat("Total Loan Volume:", f"&#8358;{cs['total_volume']:,.2f}")}{_divider()}
      {_stat("Created by Admin:", f"{cs['admin_loans']} loan{'s' if cs['admin_loans']!=1 else ''}")}
      {_bullet(f"Volume: <strong>&#8358;{cs['admin_volume']:,.2f}</strong>")}{_divider()}
      {_stat("Created on Merchant App:", f"{cs['app_loans']} loan{'s' if cs['app_loans']!=1 else ''}")}
      {_bullet(f"Volume: <strong>&#8358;{cs['app_volume']:,.2f}</strong>")}
    </div>

    <p style="margin:24px 0 10px;font-size:17px;font-weight:bold;color:#235D3B;
               border-bottom:2px solid #D9ECD4;padding-bottom:4px;">Comparison with Previous Week</p>

    <div style="background:#FFF6C4;border-left:4px solid #CFBA49;padding:16px;">
      {_comp("Total Loan Count",
             cs['total_loans'],   ps['total_loans'],
             cs['total_loans'],   ps['total_loans'])}{_ydivider()}
      {_comp("Total Loan Volume",
             f"&#8358;{cs['total_volume']:,.2f}", f"&#8358;{ps['total_volume']:,.2f}",
             cs['total_volume'],  ps['total_volume'], is_currency=True)}{_ydivider()}
      <p style="margin:4px 0;font-size:14px;font-family:Arial,sans-serif;">
        <strong>Merchants With Loans</strong><br>
        This Week: <strong>{cs['merchants_with_loans']}</strong>
        &nbsp;|&nbsp; Last Week: {ps['merchants_with_loans']}
      </p>{_ydivider()}
      <p style="margin:4px 0;font-size:14px;font-family:Arial,sans-serif;">
        <strong>Admin vs App — Loan Count</strong><br>
        This Week — Admin: <strong>{cs['admin_loans']}</strong>
        &nbsp;|&nbsp; App: <strong>{cs['app_loans']}</strong><br>
        Last Week — Admin: {ps['admin_loans']} &nbsp;|&nbsp; App: {ps['app_loans']}
      </p>{_ydivider()}
      {_comp("Admin Loan Volume",
             f"&#8358;{cs['admin_volume']:,.2f}", f"&#8358;{ps['admin_volume']:,.2f}",
             cs['admin_volume'],  ps['admin_volume'], is_currency=True)}{_ydivider()}
      {_comp("Merchant App Loan Volume",
             f"&#8358;{cs['app_volume']:,.2f}", f"&#8358;{ps['app_volume']:,.2f}",
             cs['app_volume'],    ps['app_volume'], is_currency=True)}
    </div>

    <p style="margin:20px 0 0;font-size:14px;font-family:Arial,sans-serif;">
      Please find the full data attached as an Excel file.</p>
    <p style="margin:28px 0 0;font-size:12px;color:#888;text-align:center;font-family:Arial,sans-serif;">
      Automated report generated on {ts}</p>
  </td></tr>

</table></td></tr></table>
</body></html>"""


# ---------------------------------------------------------------------------
# 4. EXCEL BUILDER
# ---------------------------------------------------------------------------

def save_excel(df, filename):
    from openpyxl.styles import Font, PatternFill
    writer      = pd.ExcelWriter(filename, engine='openpyxl')
    df.to_excel(writer, sheet_name='Weekly Report', index=False)
    ws          = writer.sheets['Weekly Report']
    header_fill = PatternFill("solid", fgColor="235D3B")
    bold_white  = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = bold_white
        cell.fill = header_fill
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)
    writer.close()


# ---------------------------------------------------------------------------
# 5. EMAIL DISPATCH  (ZeptoMail)
# ---------------------------------------------------------------------------

def send_email(filename, subject, html_body):
    with open(filename, 'rb') as f:
        encoded = base64.b64encode(f.read()).decode('utf-8')

    payload = {
        "from":    {"address": EMAIL_SENDER, "name": "FinApp Reports"},
        "to":      [{"email_address": {"address": e.strip()}} for e in EMAIL_RECIPIENTS],
        "subject": subject,
        "htmlbody": html_body,
        "attachments": [{
            "name":      filename,
            "content":   encoded,
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }]
    }
    if EMAIL_CC and EMAIL_CC[0]:
        payload["cc"] = [{"email_address": {"address": e.strip()}} for e in EMAIL_CC]

    headers = {"Authorization": f"Zoho-enczapikey {EMAIL_API_KEY}",
               "Content-Type":  "application/json"}
    resp = requests.post(EMAIL_API_URL, json=payload, headers=headers)
    resp.raise_for_status()
    print(f"   -> Email sent (HTTP {resp.status_code})")


# ---------------------------------------------------------------------------
# 6. DATABASE
# ---------------------------------------------------------------------------

def get_conn():
    return psycopg2.connect(
        host=DB_HOST, database=DB_NAME,
        user=DB_USER, password=DB_PASS, port=DB_PORT
    )

def fetch_data(start_date, end_date):
    conn = get_conn()
    df   = pd.read_sql_query(get_report_query(start_date, end_date), conn)
    conn.close()
    return df


# ---------------------------------------------------------------------------
# 7. MAIN
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("WEEKLY MERCHANT LOAN REPORT")
    print("=" * 60)

    # Auto-calculate rolling 7-day window
    # To backfill a specific week, uncomment and set dates:
    # from datetime import date
    # curr_start = date(2026, 3, 6)
    # curr_end   = date(2026, 3, 13)
    today      = datetime.now().date()
    curr_start = today - timedelta(days=7)
    prev_start = curr_start - timedelta(days=7)
    prev_end   = curr_start

    print(f"Current : {curr_start} → {today - timedelta(days=1)}")
    print(f"Previous: {prev_start} → {prev_end - timedelta(days=1)}\n")

    curr_df = fetch_data(curr_start, today)
    prev_df = fetch_data(prev_start, prev_end)
    cs      = build_summary(curr_df)
    ps      = build_summary(prev_df)

    filename  = f"Merchant_Loan_Report_{curr_start}_to_{today}.xlsx"
    subject   = f"Weekly Merchant Loan Report - {curr_start} to {today}"
    html_body = build_email_html(cs, ps, curr_start, today)

    save_excel(curr_df, filename)
    send_email(filename, subject, html_body)
    os.remove(filename)

    print("\nREPORT SENT SUCCESSFULLY")


if __name__ == "__main__":
    main()
