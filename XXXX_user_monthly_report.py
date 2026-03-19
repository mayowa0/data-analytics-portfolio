"""
Automated Monthly Individual User Report
------------------------------------------
Runs on the 1st of every month. Covers the previous full calendar month.
Excludes merchant users entirely.
Tracks: signups by tier, loan disbursements (app vs admin), conversion rates,
application funnel, and decline rates.

Author : Mayowa Alamutu
Stack  : Python · psycopg2 · pandas · openpyxl · ZeptoMail API
"""

import psycopg2
import pandas as pd
import requests
import base64
from datetime import datetime, timedelta, date
import os
from dotenv import load_dotenv

load_dotenv()

DB_HOST = os.getenv('DB_HOST')
DB_NAME = os.getenv('DB_NAME')
DB_USER = os.getenv('DB_USER')
DB_PASS = os.getenv('DB_PASS')
DB_PORT = os.getenv('DB_PORT', '5432')

EMAIL_API_KEY    = os.getenv('EMAIL_ZEPTOMAIL_API_KEY')
EMAIL_API_URL    = os.getenv('EMAIL_ZEPTOMAIL_API_URL')
EMAIL_SENDER     = os.getenv('EMAIL_SENDER')
EMAIL_RECIPIENTS = os.getenv('EMAIL_RECIPIENTS').split(',')
EMAIL_CC         = os.getenv('EMAIL_CC_RECIPIENTS', '').split(',') if os.getenv('EMAIL_CC_RECIPIENTS') else []


# ---------------------------------------------------------------------------
# 1. DATE HELPERS
# ---------------------------------------------------------------------------

def month_range(year, month):
    """Return (start, end) where end is exclusive (first day of next month)."""
    start = date(year, month, 1)
    end   = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    return start, end

def month_label(year, month):
    return datetime(year, month, 1).strftime('%B %Y')


# ---------------------------------------------------------------------------
# 2. SQL QUERIES  (merchant users excluded in all queries)
# ---------------------------------------------------------------------------

def q_new_users(s, e):
    return f"""
    SELECT u.user_id, u.tier,
           l.amount_requested AS amount_disbursed,
           CASE WHEN l.is_created_by_admin = true  THEN 'Admin Created'
                WHEN l.is_created_by_admin = false THEN 'App Created'
                ELSE 'No Disbursement' END AS disbursement_source,
           TRUE AS is_new_user
    FROM users u
    LEFT JOIN personal_loans l
          ON u.user_id = l.user_id
         AND l.is_loan_disbursed = true
         AND l.loan_disbursed_at >= '{s}' AND l.loan_disbursed_at < '{e}'
         AND l.loan_id NOT IN (SELECT loan_id FROM merchant_user_loans)
    WHERE u.created_at >= '{s}' AND u.created_at < '{e}'
      AND u.user_id NOT IN (SELECT user_id FROM merchant_users)
      AND u.is_deleted = false
    ORDER BY u.created_at;
    """

def q_existing_disbursements(s, e):
    return f"""
    SELECT u.user_id, u.tier,
           l.amount_requested AS amount_disbursed,
           CASE WHEN l.is_created_by_admin = true  THEN 'Admin Created'
                ELSE 'App Created' END AS disbursement_source,
           CASE WHEN u.created_at >= '{s}' AND u.created_at < '{e}'
                THEN TRUE ELSE FALSE END AS is_new_user
    FROM personal_loans l
    JOIN users u ON u.user_id = l.user_id
    WHERE l.is_loan_disbursed = true
      AND l.loan_disbursed_at >= '{s}' AND l.loan_disbursed_at < '{e}'
      AND l.loan_id NOT IN (SELECT loan_id FROM merchant_user_loans)
      AND u.user_id NOT IN (SELECT user_id FROM merchant_users)
      AND u.is_deleted = false
    ORDER BY l.loan_disbursed_at;
    """

def q_all_applications(s, e):
    """All loan applications in period (multiple per user allowed)."""
    return f"""
    SELECT pl.user_id, pl.amount_requested, u.tier
    FROM personal_loans pl
    JOIN users u ON u.user_id = pl.user_id
    WHERE u.is_deleted = false
      AND pl.created_at >= '{s}' AND pl.created_at < '{e}'
      AND pl.loan_id NOT IN (SELECT loan_id FROM merchant_user_loans);
    """

def q_distinct_applications(s, e):
    """One record per user — most recent application."""
    return f"""
    SELECT DISTINCT ON (pl.user_id)
           pl.user_id, pl.amount_requested, u.tier
    FROM personal_loans pl
    JOIN users u ON u.user_id = pl.user_id
    WHERE u.is_deleted = false
      AND pl.created_at >= '{s}' AND pl.created_at < '{e}'
      AND pl.loan_id NOT IN (SELECT loan_id FROM merchant_user_loans)
    ORDER BY pl.user_id, pl.created_at DESC;
    """

def q_distinct_declined(s, e):
    """One declined record per user."""
    return f"""
    SELECT DISTINCT ON (pl.user_id)
           pl.user_id, u.tier
    FROM personal_loans pl
    JOIN users u ON u.user_id = pl.user_id
    WHERE pl.status = 'declined'
      AND u.is_deleted = false
      AND pl.created_at >= '{s}' AND pl.created_at < '{e}'
      AND pl.loan_id NOT IN (SELECT loan_id FROM merchant_user_loans)
    ORDER BY pl.user_id, pl.created_at DESC;
    """


# ---------------------------------------------------------------------------
# 3. SUMMARY CALCULATION
# ---------------------------------------------------------------------------

def pct_change(curr, prev):
    if prev == 0:
        return 100.0 if curr > 0 else 0.0
    return ((curr - prev) / prev) * 100


def build_summary(df, apps_df, distinct_df, declined_df):
    zero_tiers = {'0': 0, '1': 0, '2': 0}

    if df.empty:
        return {k: (0.0 if 'rate' in k or 'amount' in k or 'volume' in k else
                    zero_tiers if 'tier' in k else 0)
                for k in ['total_new_users','tier_signups','total_disbursed',
                           'amount_app','amount_admin','users_disbursed',
                           'new_disbursed','old_disbursed','app_disbursed',
                           'admin_disbursed','tier_disbursed','tier_new','tier_old',
                           'conversion_rate','total_applications',
                           'distinct_applications','distinct_declined','decline_rate']}

    df = df.copy()
    df['tier']        = df['tier'].astype(str)
    df['is_new_user'] = df['is_new_user'].astype(bool)

    new_df   = df[df['is_new_user']]
    disb_df  = df[df['disbursement_source'] != 'No Disbursement'].copy()
    disb_df['amount_disbursed'] = pd.to_numeric(disb_df['amount_disbursed'], errors='coerce').fillna(0)

    app_df   = disb_df[disb_df['disbursement_source'] == 'App Created']
    admin_df = disb_df[disb_df['disbursement_source'] == 'Admin Created']
    new_d    = disb_df[disb_df['is_new_user']]
    old_d    = disb_df[~disb_df['is_new_user']]

    total_new   = len(new_df)
    total_disb  = float(disb_df['amount_disbursed'].sum())
    conv_rate   = (len(new_d) / total_new * 100) if total_new > 0 else 0.0

    d_apps = len(distinct_df)
    d_decl = len(declined_df)

    return {
        'total_new_users':      total_new,
        'tier_signups':         {str(t): int((new_df['tier'] == str(t)).sum()) for t in range(3)},
        'total_disbursed':      total_disb,
        'amount_app':           float(app_df['amount_disbursed'].sum()),
        'amount_admin':         float(admin_df['amount_disbursed'].sum()),
        'users_disbursed':      len(disb_df),
        'new_disbursed':        len(new_d),
        'old_disbursed':        len(old_d),
        'app_disbursed':        len(app_df),
        'admin_disbursed':      len(admin_df),
        'tier_disbursed':       {str(t): int((disb_df['tier'] == str(t)).sum()) for t in range(3)},
        'tier_new':             {str(t): int((new_d['tier']   == str(t)).sum()) for t in range(3)},
        'tier_old':             {str(t): int((old_d['tier']   == str(t)).sum()) for t in range(3)},
        'conversion_rate':      conv_rate,
        'total_applications':   len(apps_df),
        'distinct_applications': d_apps,
        'distinct_declined':    d_decl,
        'decline_rate':         (d_decl / d_apps * 100) if d_apps > 0 else 0.0,
    }


# ---------------------------------------------------------------------------
# 4. EMAIL HTML BUILDER
# ---------------------------------------------------------------------------

def _stat(label, value):
    return (f'<p style="margin:4px 0;font-size:15px;font-family:Arial,sans-serif;">'
            f'<strong style="color:#1A201D;">{label}</strong>'
            f'<strong style="color:#2A8851;">&nbsp;{value}</strong></p>')

def _bullet(text):
    return (f'<p style="margin:2px 0 2px 16px;font-size:13px;'
            f'font-family:Arial,sans-serif;color:#333;">&#8226;&nbsp;{text}</p>')

def _divider():
    return '<hr style="border:none;border-top:1px solid #c2d9bb;margin:8px 0;">'

def _ydivider():
    return '<hr style="border:none;border-top:1px solid #e8d87a;margin:8px 0;">'

def _comp(title, this_val, last_val, curr, prev, is_currency=False, period='Month'):
    change = curr - prev
    pct    = pct_change(curr, prev)
    color  = "#2A8851" if change >= 0 else "#E46159"
    lbl    = "Growth"  if change >= 0 else "Decline"
    fmt    = f"&#8358;{change:+,.2f}" if is_currency else f"{change:+,}"
    return (f'<p style="margin:4px 0;font-size:14px;font-family:Arial,sans-serif;">'
            f'<strong>{title}</strong><br>'
            f'This {period}: <strong>{this_val}</strong>&nbsp;|&nbsp;Last {period}: {last_val}<br>'
            f'{lbl}: <strong style="color:{color};">{fmt} ({pct:+.1f}%)</strong></p>')


def build_email_html(cs, ps, curr_label, prev_label, start_date, end_date):
    period = (f"{curr_label} "
              f"({start_date.strftime('%d %b')} - "
              f"{(end_date - timedelta(days=1)).strftime('%d %b %Y')})")
    ts     = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Conversion rate comparison
    cr_diff  = cs['conversion_rate'] - ps['conversion_rate']
    cr_color = "#2A8851" if cr_diff >= 0 else "#E46159"
    cr_label = "Growth" if cr_diff >= 0 else "Decline"

    # Decline rate comparison
    dr_diff  = cs['decline_rate'] - ps['decline_rate']
    dr_color = "#2A8851" if dr_diff >= 0 else "#E46159"
    dr_label = "Growth" if dr_diff >= 0 else "Decline"

    summary_html = f"""
      {_stat("New Users Created:", cs['total_new_users'])}
      {_bullet(f"Tier 0: <strong>{cs['tier_signups']['0']}</strong>")}
      {_bullet(f"Tier 1: <strong>{cs['tier_signups']['1']}</strong>")}
      {_bullet(f"Tier 2: <strong>{cs['tier_signups']['2']}</strong>")}
      {_divider()}
      {_stat("Total Amount Disbursed:", f"&#8358;{cs['total_disbursed']:,.2f}")}
      {_bullet(f"From App: <strong>&#8358;{cs['amount_app']:,.2f}</strong>")}
      {_bullet(f"From Admin: <strong>&#8358;{cs['amount_admin']:,.2f}</strong>")}
      {_divider()}
      {_stat("Total Users Disbursed To:", cs['users_disbursed'])}
      {_bullet(f"New Users: <strong>{cs['new_disbursed']}</strong>")}
      {_bullet(f"Existing Users: <strong>{cs['old_disbursed']}</strong>")}
      {_bullet(f"Via App: <strong>{cs['app_disbursed']}</strong>")}
      {_bullet(f"Via Admin: <strong>{cs['admin_disbursed']}</strong>")}
      {_bullet(f"Tier 0: <strong>{cs['tier_disbursed']['0']}</strong> (New: {cs['tier_new']['0']} | Existing: {cs['tier_old']['0']})")}
      {_bullet(f"Tier 1: <strong>{cs['tier_disbursed']['1']}</strong> (New: {cs['tier_new']['1']} | Existing: {cs['tier_old']['1']})")}
      {_bullet(f"Tier 2: <strong>{cs['tier_disbursed']['2']}</strong> (New: {cs['tier_new']['2']} | Existing: {cs['tier_old']['2']})")}
      {_divider()}
      {_stat("New User Conversion Rate:", f"{cs['conversion_rate']:.1f}%")}
      {_bullet(f"{cs['new_disbursed']} of {cs['total_new_users']} new users received a loan this month")}
      {_divider()}
      {_stat("Loan Applications:", cs['total_applications'])}
      {_bullet(f"Distinct Applicants: <strong>{cs['distinct_applications']}</strong>")}
      {_bullet(f"Distinct Declined: <strong>{cs['distinct_declined']}</strong>")}
      {_bullet(f"Decline Rate: <strong>{cs['decline_rate']:.1f}%</strong>")}
    """

    comp_html = f"""
      {_comp("New Users Created",
             cs['total_new_users'], ps['total_new_users'],
             cs['total_new_users'], ps['total_new_users'])}{_ydivider()}
      {_comp("Total Amount Disbursed",
             f"&#8358;{cs['total_disbursed']:,.2f}", f"&#8358;{ps['total_disbursed']:,.2f}",
             cs['total_disbursed'], ps['total_disbursed'], is_currency=True)}{_ydivider()}
      {_comp("Total Users Disbursed To",
             f"{cs['users_disbursed']} (New: {cs['new_disbursed']} | Existing: {cs['old_disbursed']})",
             f"{ps['users_disbursed']} (New: {ps['new_disbursed']} | Existing: {ps['old_disbursed']})",
             cs['users_disbursed'], ps['users_disbursed'])}{_ydivider()}
      <p style="margin:4px 0;font-size:14px;font-family:Arial,sans-serif;">
        <strong>New User Conversion Rate</strong><br>
        This Month: <strong>{cs['conversion_rate']:.1f}%</strong>
        &nbsp;|&nbsp; Last Month: {ps['conversion_rate']:.1f}%<br>
        {cr_label}: <strong style="color:{cr_color};">{cr_diff:+.1f}%</strong>
      </p>{_ydivider()}
      {_comp("Total Loan Applications",
             cs['total_applications'], ps['total_applications'],
             cs['total_applications'], ps['total_applications'])}{_ydivider()}
      {_comp("Distinct Applicants",
             cs['distinct_applications'], ps['distinct_applications'],
             cs['distinct_applications'], ps['distinct_applications'])}{_ydivider()}
      {_comp("Distinct Declined Applicants",
             cs['distinct_declined'], ps['distinct_declined'],
             cs['distinct_declined'], ps['distinct_declined'])}{_ydivider()}
      <p style="margin:4px 0;font-size:14px;font-family:Arial,sans-serif;">
        <strong>Decline Rate</strong><br>
        This Month: <strong>{cs['decline_rate']:.1f}%</strong>
        &nbsp;|&nbsp; Last Month: {ps['decline_rate']:.1f}%<br>
        {dr_label}: <strong style="color:{dr_color};">{dr_diff:+.1f}%</strong>
      </p>
    """

    return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f5f5f5;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f5f5f5;">
<tr><td align="center" style="padding:20px 10px;">
<table width="620" cellpadding="0" cellspacing="0" border="0" style="background:#fff;max-width:620px;">
  <tr><td align="center" bgcolor="#235D3B" style="background:#235D3B;padding:24px 20px;">
    <p style="margin:0 0 6px;font-size:22px;font-weight:bold;color:#fff;">Monthly SeedFi User Report</p>
    <p style="margin:0;font-size:13px;color:#ccffcc;">Period: {period}</p>
  </td></tr>
  <tr><td style="padding:24px 20px;">
    <p style="margin:0 0 10px;font-size:17px;font-weight:bold;color:#235D3B;
               border-bottom:2px solid #D9ECD4;padding-bottom:4px;">Summary</p>
    <div style="background:#D9ECD4;border-left:4px solid #2A8851;padding:16px;">{summary_html}</div>
    <p style="margin:24px 0 10px;font-size:17px;font-weight:bold;color:#235D3B;
               border-bottom:2px solid #D9ECD4;padding-bottom:4px;">
      Comparison with Previous Month ({prev_label})</p>
    <div style="background:#FFF6C4;border-left:4px solid #CFBA49;padding:16px;">{comp_html}</div>
    <p style="margin:20px 0 0;font-size:14px;">Please find the full data attached as an Excel file.</p>
    <p style="margin:28px 0 0;font-size:12px;color:#888;text-align:center;">
      Automated report generated on {ts}</p>
  </td></tr>
</table></td></tr></table>
</body></html>"""


# ---------------------------------------------------------------------------
# 5. EXCEL BUILDER  (4 sheets)
# ---------------------------------------------------------------------------

def save_excel(df, apps_df, distinct_df, declined_df, filename):
    from openpyxl.styles import Font, PatternFill
    GREEN  = PatternFill("solid", fgColor="D9ECD4")
    HEADER = PatternFill("solid", fgColor="235D3B")
    RED    = PatternFill("solid", fgColor="FADADD")
    BW     = Font(bold=True, color="FFFFFF")
    BOLD   = Font(bold=True)

    def style_header(ws):
        for c in ws[1]: c.font = BW; c.fill = HEADER
    def auto_width(ws):
        for col in ws.columns:
            mx = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 50)

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Sheet 1 — All Users
    display = df.drop(columns=['user_id','is_new_user'], errors='ignore').rename(columns={
        'tier': 'Tier', 'amount_disbursed': 'Amount Disbursed (N)',
        'disbursement_source': 'Disbursement Source'})
    display.to_excel(writer, sheet_name='All Users', index=False)
    ws = writer.sheets['All Users']
    style_header(ws); auto_width(ws)

    # Sheet 2 — Tier Summary
    disb = df[df['disbursement_source'] != 'No Disbursement'].copy()
    disb['amount_disbursed'] = pd.to_numeric(disb['amount_disbursed'], errors='coerce').fillna(0)
    new_d  = disb[disb['is_new_user'] == True]
    old_d  = disb[disb['is_new_user'] == False]
    new_all= df[df['is_new_user'] == True]
    rows   = []
    for t in ['0','1','2']:
        td = disb[disb['tier'] == t]
        rows.append({'Tier': f'Tier {t}',
                     'New Signups': int((new_all['tier']==t).sum()),
                     'Users Disbursed': len(td),
                     'New Disbursed': len(new_d[new_d['tier']==t]),
                     'Existing Disbursed': len(old_d[old_d['tier']==t]),
                     'Amount (N)': float(td['amount_disbursed'].sum())})
    rows.append({'Tier':'TOTAL','New Signups':len(new_all),
                 'Users Disbursed':len(disb),'New Disbursed':len(new_d),
                 'Existing Disbursed':len(old_d),'Amount (N)':float(disb['amount_disbursed'].sum())})
    pd.DataFrame(rows).to_excel(writer, sheet_name='Tier Summary', index=False)
    ws2 = writer.sheets['Tier Summary']
    style_header(ws2); auto_width(ws2)
    for c in ws2[len(rows)+1]: c.font=BOLD; c.fill=GREEN

    # Sheet 3 — Applications
    apps_df.drop(columns=['user_id'], errors='ignore').to_excel(
        writer, sheet_name='Loan Applications', index=False)
    style_header(writer.sheets['Loan Applications'])

    # Sheet 4 — Declined
    declined_df.drop(columns=['user_id'], errors='ignore').to_excel(
        writer, sheet_name='Declined Applicants', index=False)
    ws4 = writer.sheets['Declined Applicants']
    for c in ws4[1]: c.font=BW; c.fill=PatternFill("solid", fgColor="C0392B")

    writer.close()


# ---------------------------------------------------------------------------
# 6. EMAIL DISPATCH
# ---------------------------------------------------------------------------

def send_email(filename, subject, html_body):
    with open(filename, 'rb') as f:
        encoded = base64.b64encode(f.read()).decode('utf-8')
    payload = {
        "from":    {"address": EMAIL_SENDER, "name": "FinApp Reports"},
        "to":      [{"email_address": {"address": e.strip()}} for e in EMAIL_RECIPIENTS],
        "subject": subject, "htmlbody": html_body,
        "attachments": [{"name": filename, "content": encoded,
                          "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}]
    }
    if EMAIL_CC and EMAIL_CC[0]:
        payload["cc"] = [{"email_address": {"address": e.strip()}} for e in EMAIL_CC]
    headers = {"Authorization": f"Zoho-enczapikey {EMAIL_API_KEY}", "Content-Type": "application/json"}
    resp = requests.post(EMAIL_API_URL, json=payload, headers=headers)
    resp.raise_for_status()
    print(f"   -> Email sent (HTTP {resp.status_code})")


# ---------------------------------------------------------------------------
# 7. DATABASE
# ---------------------------------------------------------------------------

def get_conn():
    return psycopg2.connect(host=DB_HOST, database=DB_NAME,
                             user=DB_USER, password=DB_PASS, port=DB_PORT)

def strip_tz(df):
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            try:
                if df[col].dt.tz: df[col] = df[col].dt.tz_localize(None)
            except Exception: pass
    return df

def fetch_data(s, e):
    conn = get_conn()
    new_df    = strip_tz(pd.read_sql_query(q_new_users(s, e), conn))
    all_df    = strip_tz(pd.read_sql_query(q_existing_disbursements(s, e), conn))
    apps_df   = strip_tz(pd.read_sql_query(q_all_applications(s, e), conn))
    dist_df   = strip_tz(pd.read_sql_query(q_distinct_applications(s, e), conn))
    decl_df   = strip_tz(pd.read_sql_query(q_distinct_declined(s, e), conn))
    conn.close()
    # Merge: avoid double-counting new users in existing disbursements
    new_ids = set(new_df['user_id'].tolist()) if 'user_id' in new_df.columns else set()
    old_df  = all_df[~all_df['user_id'].isin(new_ids)]
    df      = pd.concat([new_df, old_df], ignore_index=True)
    df['tier'] = df['tier'].astype(str)
    df['is_new_user'] = df['is_new_user'].astype(bool)
    return df, apps_df, dist_df, decl_df


# ---------------------------------------------------------------------------
# 8. MAIN
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("MONTHLY INDIVIDUAL USER REPORT")
    print("=" * 60)

    # Auto-calculate: run on 1st, covers previous full month
    # To backfill: uncomment and set year/month:
    # report_year, report_month = 2026, 2
    today        = datetime.now().date()
    report_year  = today.year  if today.month > 1 else today.year - 1
    report_month = today.month - 1 if today.month > 1 else 12
    prev_year    = report_year  if report_month > 1 else report_year - 1
    prev_month   = report_month - 1 if report_month > 1 else 12

    cs_start, cs_end = month_range(report_year, report_month)
    ps_start, ps_end = month_range(prev_year, prev_month)
    cl = month_label(report_year, report_month)
    pl = month_label(prev_year, prev_month)

    print(f"Current : {cl}  ({cs_start} → {cs_end - timedelta(days=1)})")
    print(f"Previous: {pl}  ({ps_start} → {ps_end - timedelta(days=1)})\n")

    curr_df, curr_apps, curr_dist, curr_decl = fetch_data(cs_start, cs_end)
    prev_df, prev_apps, prev_dist, prev_decl = fetch_data(ps_start, ps_end)
    cs = build_summary(curr_df, curr_apps, curr_dist, curr_decl)
    ps = build_summary(prev_df, prev_apps, prev_dist, prev_decl)

    filename  = f"User_Monthly_Report_{cl.replace(' ', '_')}.xlsx"
    subject   = f"Monthly Individual User Report - {cl}"
    html_body = build_email_html(cs, ps, cl, pl, cs_start, cs_end)

    save_excel(curr_df, curr_apps, curr_dist, curr_decl, filename)
    send_email(filename, subject, html_body)
    os.remove(filename)

    print("\nREPORT SENT SUCCESSFULLY")


if __name__ == "__main__":
    main()
